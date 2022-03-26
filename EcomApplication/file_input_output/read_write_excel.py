import random
import openpyxl
import xlsxwriter

cities = ['Pune', 'Mumbai', 'Banglore', 'Hydrabad', 'Gurgaon', 'Delhi']


class Student:
    def __init__(self, rollno, sname, address, marks):
        self.student_rolllno = int(rollno)
        self.student_name = str(sname)
        self.student_address = str(address)
        self.student_marks = float(marks)
        Student.count = Student.count + 1

    def __str__(self):
        return f'''\n{self.__dict__}'''

    def __repr__(self):
        return str(self)

    count = 1

    @classmethod
    def get_student_info(cls):
        return cls(rollno=Student.count,
                   sname='Student' + str(Student.count),
                   address=cities[random.randint(0, len(cities) - 1)],
                   marks=round(random.randint(1, 600) / 6, 2)
                   )


studList = []
for i in range(9):
    studList.append(Student.get_student_info())
print('------------------------------------------------------------------------------------------------------------')
print(studList)
print('------------------------------------------------------------------------------------------------------------')

file_path = "F:\\pythonProject\\EcomApplication\\file_input_output\\data_files\\"


def write_data_into_excel(studList):
    workbook = xlsxwriter.Workbook(file_path + 'Student.xlsx')
    worksheet = workbook.add_worksheet()
    count = 0
    for stud in studList:
        count = count + 1
        worksheet.write('A' + str(count), stud.student_rolllno)
        worksheet.write('B' + str(count), stud.student_name)
        worksheet.write('C' + str(count), stud.student_address)
        worksheet.write('D' + str(count), stud.student_marks)
    workbook.close()

    print("Student Data are added in Student.xlsx file.....")


def read_data_from_excel():
    print("Read Data from Student.xlsx file.....")
    slist = []
    path = file_path + 'Student.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column

    for i in range(1, m_row + 1):
        rowline = []
        for j in range(1, m_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            rowline.append(str(cell_obj.value))

        slist.append(Student(rollno=rowline[0], sname=rowline[1], address=rowline[2], marks=rowline[3]))
    print(slist)


# call function
write_data_into_excel(studList)
read_data_from_excel()
