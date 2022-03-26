import random
import openpyxl

class Employee:
    def __init__(self, eid, ename, salary):
        self.employee_id = eid
        self.employee_name = ename
        self.employee_salary = salary
        Employee.count = Employee.count + 1

    def __str__(self):
        return f''''\n{self.__dict__}'''

    def __repr__(self):
        return str(self)

    count = 1

    @classmethod
    def get_employeee_info(cls):
        return cls(eid=Employee.count,
                   ename="Employee" + str(Employee.count),
                   salary=round(random.randint(10000, 20000) / 3, 2)
                   )

employeeList = []
for i in range(10):
    employeeList.append(Employee.get_employeee_info())
print('------------------------------------------------------------------------------------------------------------')
print(employeeList)
print('------------------------------------------------------------------------------------------------------------')

def write_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet('emp_data')

    #creating Headers
    headerNames = employeeList[0].__dict__.keys()

    count = 65
    for header in headerNames:
        sheet[chr(count)+str(1)] = header
        count = count + 1

    rownumber = 2
    for emp in employeeList:
        row = str(rownumber)
        sheet["A" + row] = emp.employee_id
        sheet["B" + row] = emp.employee_name
        sheet["C" + row] = emp.employee_salary
        rownumber = rownumber + 1

    workbook.save("F:\\pythonProject\\EcomApplication\\file_input_output\\data_files\\Employee.xlsx")


def read_excel():
    print("Read Data from Employee.xlsx file.....")
    empList = []
    path = "F:\\pythonProject\\EcomApplication\\file_input_output\\data_files\\Employee.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj['emp_data']

    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column

    for i in range(1, m_row + 1):
        rowline = []
        for j in range(1, m_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            rowline.append(str(cell_obj.value))
        empList.append(Employee(eid = rowline[0], ename = rowline[1], salary = rowline[2] ))
    print(empList)

write_excel()
read_excel()